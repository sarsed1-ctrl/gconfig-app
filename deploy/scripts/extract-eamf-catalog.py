#!/usr/bin/env python3
"""Extract EAMF materials + kants + countertops from ALL MATERIALS.xlsx -> assets/eamf-catalog.json"""
import json
import os
import re
import shutil
import sys
import tempfile
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl

SRC_CANDIDATES = [
    Path(r"C:\Users\georgi\OneDrive - AM furnitura\Desktop\ALL MATERIALS.xlsx"),
    Path(r"C:\Users\georgi\Downloads\ALL MATERIALS.xlsx"),
]
ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "assets" / "eamf-catalog.json"
DEPLOY_OUT = ROOT / "deploy" / "assets" / "eamf-catalog.json"

# BEDS mode MDF carcass — not in ALL MATERIALS.xlsx plate column; always inject.
BED_MDF_CARCASS_ARTICLES = [
    ("77.124.D.16", 16),
    ("77.124.D.18", 18),
    ("77.124.D.19", 19),
    ("77.124.D.25", 25),
    ("77.124.D.28", 28),
]

# Used when Excel is unavailable (manual catalog patch).
FALLBACK_COUNTERTOP_ARTICLES = [
    "76.0026.12",
    "76.0077.12",
    "76.0164.MM.30.12",
]

CT_EXCLUDED_COUNTERTOPS = frozenset({"76.0077.8"})
CT_HPL12_ARTICLES = frozenset({"76.F206.C.920", "76.F274.C.920"})


def is_hpl12_countertop(code: str, parts: list[str]) -> bool:
    if code in CT_HPL12_ARTICLES:
        return True
    return has_trailing_thickness(parts) and parts[-1] == "12"


def resolve_source_path() -> Path:
    for candidate in SRC_CANDIDATES:
        if candidate.is_file():
            try:
                with candidate.open("rb"):
                    pass
                return candidate
            except OSError:
                pass
            tmp = Path(tempfile.gettempdir()) / "ALL_MATERIALS_copy.xlsx"
            try:
                shutil.copy2(candidate, tmp)
                return tmp
            except OSError:
                continue
    return SRC_CANDIDATES[0]


def detect_column_indices(header_row: tuple) -> dict[str, int | None]:
    cols: dict[str, int | None] = {
        "plate": 0,
        "countertop": None,
        "edge": 3,
        "hdf": 4,
    }
    for idx, cell in enumerate(header_row):
        text = str(cell or "").strip().lower()
        if not text:
            continue
        if "73" in text or "плит" in text:
            cols["plate"] = idx
        elif "76" in text or "столешниц" in text:
            cols["countertop"] = idx
        elif "74" in text or "кант" in text:
            cols["edge"] = idx
        elif "hdf" in text or "задн" in text:
            cols["hdf"] = idx
    return cols


def parse_plate(code: str) -> dict:
    code = str(code).strip()
    parts = code.split(".")
    thick = None
    for p in reversed(parts):
        if p.isdigit():
            thick = int(p)
            break
    decor = None
    for p in parts:
        if re.match(r"^[A-Z]{1,2}\d{3,4}$", p):
            decor = p
            break
    label = code
    feel_wood = bool(re.search(r"\.TM\d+\.", code, re.I))
    return {
        "code": code,
        "article": code,
        "decor": decor,
        "thick": thick,
        "thickness": thick,
        "postform": ".P." in code,
        "feelWood": feel_wood,
        "label": label,
    }


CT_DEPTH_MARKERS = frozenset({"60", "65", "90", "92"})
CT_TRAILING_THICKNESS = frozenset({"8", "12", "13", "20"})


def has_trailing_thickness(parts: list[str]) -> bool:
    return len(parts) >= 3 and parts[-1] in CT_TRAILING_THICKNESS


def is_ldsp_38_countertop(parts: list[str]) -> bool:
    """76.* LDSP countertops: suffix .1.5 or depth markers 60/65/90/92 (not HPL/other trailing thickness)."""
    if not parts or parts[0] != "76" or len(parts) < 3:
        return False
    if has_trailing_thickness(parts):
        return False
    if len(parts) >= 2 and parts[-2] == "1" and parts[-1] == "5":
        return True
    return any(p in CT_DEPTH_MARKERS for p in parts[1:])


def get_countertop_depth(parts: list[str]) -> int | None:
    if len(parts) >= 4 and parts[-1] in CT_TRAILING_THICKNESS and parts[-2] in CT_DEPTH_MARKERS:
        return int(parts[-2])
    if len(parts) >= 4 and parts[-2] == "1" and parts[-1] == "5":
        if parts[-3] in CT_DEPTH_MARKERS:
            return int(parts[-3])
    for p in parts[1:]:
        if p in CT_DEPTH_MARKERS:
            return int(p)
    return None


def parse_countertop(code: str) -> dict:
    code = str(code).strip()
    parts = code.split(".")
    hpl12 = is_hpl12_countertop(code, parts)
    hdf13 = has_trailing_thickness(parts) and parts[-1] == "13"
    ldsp20 = has_trailing_thickness(parts) and parts[-1] == "20"
    ldsp38 = is_ldsp_38_countertop(parts)
    depth = get_countertop_depth(parts)
    thick = None
    thick_idx = None
    decor = None

    if hpl12:
        thick = 12
        if code in CT_HPL12_ARTICLES:
            decor = ".".join(parts[1:-2]) if len(parts) > 3 else parts[1]
        elif depth is not None and len(parts) >= 4 and parts[-2] == str(depth):
            decor = ".".join(parts[1:-2]) if len(parts) > 3 else parts[1]
        else:
            decor = ".".join(parts[1:-1])
    elif hdf13:
        thick = 13
        if depth is not None and len(parts) >= 4 and parts[-2] == str(depth):
            decor = ".".join(parts[1:-2]) if len(parts) > 3 else parts[1]
        else:
            decor = ".".join(parts[1:-1])
    elif ldsp20:
        thick = 20
        if depth is not None and len(parts) >= 4 and parts[-2] == str(depth):
            decor = ".".join(parts[1:-2]) if len(parts) > 3 else parts[1]
        else:
            decor = ".".join(parts[1:-1])
    elif ldsp38:
        thick = 38
        if depth is not None:
            depth_idx = next(i for i, p in enumerate(parts) if i > 0 and p == str(depth))
            decor = ".".join(parts[1:depth_idx]) if depth_idx > 1 else (parts[1] if len(parts) > 1 else None)
        elif len(parts) >= 4 and parts[-2] == "1" and parts[-1] == "5":
            decor = ".".join(parts[1:-3]) if len(parts) > 4 else parts[1]
    else:
        for i in range(len(parts) - 1, 0, -1):
            if parts[i].isdigit():
                thick = int(parts[i])
                thick_idx = i
                break
        if len(parts) >= 3 and parts[0] == "76" and thick_idx is not None and thick_idx > 1:
            decor = ".".join(parts[1:thick_idx])

    return {
        "code": code,
        "article": code,
        "decor": decor,
        "thick": thick,
        "thickness": thick,
        "depth": depth,
        "ldsp38": ldsp38,
        "hpl12": hpl12,
        "hdf13": hdf13,
        "ldsp20": ldsp20,
        "label": code,
    }


def parse_hdf_back(code: str) -> dict:
    code = str(code).strip()
    parts = code.split(".")
    decor = None
    for p in parts[1:]:
        if re.match(r"^[A-Z]{1,2}\d{3,4}$", p):
            decor = p
            break
    thick = None
    if len(parts) > 2 and re.fullmatch(r"\d+\.?\d*", parts[-1] or ""):
        try:
            thick = float(parts[-1]) if "." in parts[-1] else int(parts[-1])
        except ValueError:
            thick = None
    return {
        "code": code,
        "article": code,
        "decor": decor,
        "thick": thick,
        "label": code,
    }


def parse_edge(code: str) -> dict:
    code = str(code).strip()
    parts = code.split(".")
    nums = [int(p) for p in parts if re.fullmatch(r"\d+", p)]
    thick = width = None
    if len(nums) >= 2:
        thick, width = nums[-2], nums[-1]
    decor = parts[1] if len(parts) > 1 else None
    return {
        "code": code,
        "article": code,
        "decor": decor,
        "thick": thick,
        "width": width,
        "label": code,
    }


def main() -> int:
    src = resolve_source_path()
    if not src.is_file():
        print("Source not found. Tried:", ", ".join(str(p) for p in SRC_CANDIDATES), file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    sheet_name = wb.sheetnames[0]
    all_rows = list(wb.active.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        print("Workbook is empty", file=sys.stderr)
        return 1

    header = all_rows[0]
    cols = detect_column_indices(header)
    rows = all_rows[1:]

    decor_thick_edges: dict[tuple, set[str]] = defaultdict(set)
    row_pair: dict[str, str] = {}
    plates_raw: list[dict] = []
    edges_set: dict[str, dict] = {}
    back_panels_set: dict[str, dict] = {}
    plate_to_back_panel: dict[str, str] = {}
    decor_to_back_panel: dict[str, str] = {}
    countertops_set: dict[str, dict] = {}

    plate_col = cols["plate"] if cols["plate"] is not None else 0
    edge_col = cols["edge"]
    hdf_col = cols["hdf"]
    ct_col = cols["countertop"]

    for r in rows:
        plate = ""
        if plate_col is not None and plate_col < len(r) and r[plate_col]:
            plate = str(r[plate_col]).strip()
        edge = ""
        if edge_col is not None and edge_col < len(r) and r[edge_col]:
            edge = str(r[edge_col]).strip()
        hdf = ""
        if hdf_col is not None and hdf_col < len(r) and r[hdf_col]:
            hdf = str(r[hdf_col]).strip()
        countertop = ""
        if ct_col is not None and ct_col < len(r) and r[ct_col]:
            countertop = str(r[ct_col]).strip()

        if countertop and countertop.startswith("76.") and countertop not in CT_EXCLUDED_COUNTERTOPS:
            ct = parse_countertop(countertop)
            countertops_set[ct["code"]] = ct

        if hdf:
            bp = parse_hdf_back(hdf)
            back_panels_set[bp["code"]] = bp
            if bp["decor"]:
                decor_to_back_panel[bp["decor"]] = bp["code"]
        if edge:
            e = parse_edge(edge)
            edges_set[e["code"]] = e
        if not plate:
            continue
        p = parse_plate(plate)
        plates_raw.append(p)
        if hdf:
            plate_to_back_panel[p["code"]] = hdf
            if p["decor"]:
                decor_to_back_panel[p["decor"]] = hdf
        if edge:
            row_pair[p["code"]] = e["code"]
            if p["decor"] and p["thick"] is not None:
                decor_thick_edges[(p["decor"], p["thick"])].add(e["code"])

    materials: list[dict] = []
    seen_m: set[str] = set()
    for p in plates_raw:
        if p["code"] in seen_m:
            continue
        seen_m.add(p["code"])
        edge_codes: set[str] = set()
        if p["code"] in row_pair:
            edge_codes.add(row_pair[p["code"]])
        if p["decor"] and p["thick"] is not None:
            edge_codes |= decor_thick_edges.get((p["decor"], p["thick"]), set())
        materials.append({**p, "edgeCodes": sorted(edge_codes)})

    for code, thick in BED_MDF_CARCASS_ARTICLES:
        if code in seen_m:
            continue
        seen_m.add(code)
        materials.append({
            "code": code,
            "article": code,
            "decor": "124",
            "thick": thick,
            "thickness": thick,
            "postform": False,
            "feelWood": False,
            "bedMdfCarcass": True,
            "label": code,
            "edgeCodes": [],
        })

    if not countertops_set:
        for code in FALLBACK_COUNTERTOP_ARTICLES:
            ct = parse_countertop(code)
            countertops_set[ct["code"]] = ct

    edges = sorted(edges_set.values(), key=lambda x: x["code"])
    materials.sort(key=lambda m: (m.get("thick") or 0, m.get("decor") or "", m["code"]))
    back_panels = sorted(back_panels_set.values(), key=lambda x: x["code"])
    countertops = sorted(countertops_set.values(), key=lambda x: (x.get("thick") or 0, x.get("decor") or "", x["code"]))

    catalog = {
        "source": src.name,
        "sheet": sheet_name,
        "generated": date.today().isoformat(),
        "counts": {
            "materials": len(materials),
            "edges": len(edges),
            "backPanels": len(back_panels),
            "countertops": len(countertops),
        },
        "materials": materials,
        "edges": edges,
        "backPanels": back_panels,
        "countertops": countertops,
        "plateToBackPanel": dict(sorted(plate_to_back_panel.items())),
        "decorToBackPanel": dict(sorted(decor_to_back_panel.items())),
    }

    for out_path in (OUT, DEPLOY_OUT):
        out_path.parent.mkdir(parents=True, exist_ok=True)
        with out_path.open("w", encoding="utf-8") as f:
            json.dump(catalog, f, ensure_ascii=False, indent=2)
        print(f"Wrote {out_path}")

    print(
        f"materials={len(materials)} edges={len(edges)} "
        f"backPanels={len(back_panels)} countertops={len(countertops)} "
        f"platePairs={len(plate_to_back_panel)}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
