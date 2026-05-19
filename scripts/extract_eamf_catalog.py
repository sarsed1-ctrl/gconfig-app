"""Extract EAMF materials and edges from tellimus-template.xlsx -> assets/eamf-catalog.json."""
from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "assets" / "tellimus-template.xlsx"
OUT = ROOT / "assets" / "eamf-catalog.json"


def norm_decor(value: str | None) -> str:
    return re.sub(r"[^A-Z0-9]", "", (value or "").upper())


def parse_bool(value) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    s = str(value).strip().upper()
    return s in ("TRUE", "1", "YES", "Y", "JAH")


def parse_num(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def main() -> None:
    wb = openpyxl.load_workbook(TEMPLATE, read_only=True, data_only=True)
    ws_mat = wb["DB_materiali"]
    ws_edge = wb["DB_malas"]

    materials = []
    for row in ws_mat.iter_rows(min_row=2, values_only=True):
        article = row[0]
        if not article:
            continue
        article = str(article).strip()
        decor = str(row[1]).strip() if row[1] is not None else ""
        grain = parse_bool(row[2])
        length = parse_num(row[3])
        thickness = parse_num(row[4])
        width = parse_num(row[5])
        default_x1 = row[6] if len(row) > 6 else None
        default_x2 = row[7] if len(row) > 7 else None
        texture = row[18] if len(row) > 18 else None  # col S (0-based 18)
        grain_dir = row[19] if len(row) > 19 else None  # col T

        materials.append(
            {
                "article": article,
                "decor": decor,
                "decorNorm": norm_decor(decor),
                "grain": grain,
                "length": length,
                "thickness": thickness,
                "width": width,
                "defaultEdgeX1": str(default_x1).strip() if default_x1 else "",
                "defaultEdgeX2": str(default_x2).strip() if default_x2 else "",
                "texture": str(texture).strip() if texture else "",
                "grainDir": str(grain_dir).strip() if grain_dir else "",
                "label": f"{article} — {decor} ({int(thickness) if thickness == int(thickness or 0) else thickness} mm)"
                if thickness is not None
                else f"{article} — {decor}",
            }
        )

    edges = []
    for row in ws_edge.iter_rows(min_row=2, values_only=True):
        article = row[0]
        if not article:
            continue
        article = str(article).strip()
        decor = str(row[1]).strip() if row[1] is not None else ""
        min_t = parse_num(row[2])
        max_t = parse_num(row[3])
        edges.append(
            {
                "article": article,
                "decor": decor,
                "decorNorm": norm_decor(decor),
                "minThickness": min_t,
                "maxThickness": max_t,
                "isUniversal": "ORG" in norm_decor(decor) or "ORG" in article.upper(),
                "label": f"{article} — {decor} ({int(min_t)}–{int(max_t)} mm)"
                if min_t is not None and max_t is not None
                else f"{article} — {decor}",
            }
        )

    wb.close()

    catalog = {
        "source": str(TEMPLATE.name),
        "generatedFrom": "DB_materiali, DB_malas",
        "materialCount": len(materials),
        "edgeCount": len(edges),
        "materials": materials,
        "edges": edges,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT}")
    print(f"  materials: {len(materials)}")
    print(f"  edges: {len(edges)}")


if __name__ == "__main__":
    main()
