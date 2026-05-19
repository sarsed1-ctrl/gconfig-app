#!/usr/bin/env python3
"""Extract EAMF materials + kants from ALL MATERIALS.xlsx -> assets/eamf-catalog.json"""
import json
import re
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl

SRC_CANDIDATES = [
    Path(r"C:\Users\georgi\OneDrive - AM furnitura\Desktop\ALL MATERIALS.xlsx"),
    Path(r"C:\Users\georgi\Downloads\ALL MATERIALS.xlsx"),
]
OUT = Path(__file__).resolve().parent.parent / "assets" / "eamf-catalog.json"


def resolve_source_path() -> Path:
    for candidate in SRC_CANDIDATES:
        if candidate.is_file():
            return candidate
    return SRC_CANDIDATES[0]


def parse_plate(code: str) -> dict:
    code = str(code).strip()
    parts = code.split(".")
    thick = None
    for p in reversed(parts):
        if p.isdigit():
            thick = int(p)
            break
    decor = None
    for p in parts:
        if re.match(r"^[A-Z]{1,2}\d{3,4}$", p):
            decor = p
            break
    label = code
    feel_wood = bool(re.search(r"\.TM\d+\.", code, re.I))
    return {
        "code": code,
        "article": code,
        "decor": decor,
        "thick": thick,
        "thickness": thick,
        "postform": ".P." in code,
        "feelWood": feel_wood,
        "label": label,
    }


def parse_edge(code: str) -> dict:
    code = str(code).strip()
    parts = code.split(".")
    nums = [int(p) for p in parts if re.fullmatch(r"\d+", p)]
    thick = width = None
    if len(nums) >= 2:
        thick, width = nums[-2], nums[-1]
    decor = parts[1] if len(parts) > 1 else None
    return {
        "code": code,
        "article": code,
        "decor": decor,
        "thick": thick,
        "width": width,
        "label": code,
    }


def main() -> int:
    src = resolve_source_path()
    if not src.is_file():
        print("Source not found. Tried:", ", ".join(str(p) for p in SRC_CANDIDATES), file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    sheet_name = wb.sheetnames[0]
    rows = list(wb.active.iter_rows(values_only=True))[1:]
    wb.close()

    decor_thick_edges: dict[tuple, set[str]] = defaultdict(set)
    row_pair: dict[str, str] = {}
    plates_raw: list[dict] = []
    edges_set: dict[str, dict] = {}

    for r in rows:
        plate = str(r[0]).strip() if r[0] else ""
        edge = str(r[3]).strip() if len(r) > 3 and r[3] else ""
        if edge:
            e = parse_edge(edge)
            edges_set[e["code"]] = e
        if not plate:
            continue
        p = parse_plate(plate)
        plates_raw.append(p)
        if edge:
            row_pair[p["code"]] = e["code"]
            if p["decor"] and p["thick"] is not None:
                decor_thick_edges[(p["decor"], p["thick"])].add(e["code"])

    materials: list[dict] = []
    seen_m: set[str] = set()
    for p in plates_raw:
        if p["code"] in seen_m:
            continue
        seen_m.add(p["code"])
        edge_codes: set[str] = set()
        if p["code"] in row_pair:
            edge_codes.add(row_pair[p["code"]])
        if p["decor"] and p["thick"] is not None:
            edge_codes |= decor_thick_edges.get((p["decor"], p["thick"]), set())
        materials.append({**p, "edgeCodes": sorted(edge_codes)})

    edges = sorted(edges_set.values(), key=lambda x: x["code"])
    materials.sort(key=lambda m: (m.get("thick") or 0, m.get("decor") or "", m["code"]))

    catalog = {
        "source": src.name,
        "sheet": sheet_name,
        "generated": date.today().isoformat(),
        "counts": {"materials": len(materials), "edges": len(edges)},
        "materials": materials,
        "edges": edges,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8") as f:
        json.dump(catalog, f, ensure_ascii=False, indent=2)

    print(f"Wrote {OUT}")
    print(f"materials={len(materials)} edges={len(edges)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
